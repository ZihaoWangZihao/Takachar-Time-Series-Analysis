import pandas as pd
import matplotlib.pyplot as plt
from datetime import datetime, timedelta

from openpyxl import load_workbook

file_path = "/Users/zihaowang/PycharmProjects/Time-Series-Analysis/Takachar-Time-Series-Analysis/Hot_Test_15/NSTPROR00006-2023-06-22 (1).xlsx"
feeding_rate_path = "/Users/zihaowang/PycharmProjects/Time-Series-Analysis/Takachar-Time-Series-Analysis/Hot_Test_15/Feeding Rate 22 nd June 2023.xlsx"

df = pd.read_excel(file_path)

col_letters = {'A': 0, 'B': 1, 'C': 2, 'D': 3, 'E': 4, 'F': 5, 'G': 6, 'H': 7, 'I': 8, 'J': 9,
               'K': 10, 'L': 11, 'M': 12, 'N': 13, 'O': 14, 'P': 15, 'Q': 16, 'R': 17, 'S': 18,
               'T': 19, 'U': 20, 'V': 21, 'W': 22, 'X': 23, 'Y': 24, 'Z': 25}

fan_cols = ["C", "D", "E", "F", "G", "H"]


def fan_data(columns):
    """
    :param columns: <list> fan columns in letter-based format
    :param file: <str> file_path
    :return: Excel file edited with all values in the Fan columns adjusted
    """

    for letter in columns:
        col_to_change = df.iloc[:, col_letters[letter]]
        for i, value in enumerate(col_to_change):
            if value[-1] == "R":
                new_value = value.split("-")
                df.iloc[i, col_letters[letter]] = int(new_value[0])  # Set the value in the current cell
            else:
                new_value = 0
                df.iloc[i, col_letters[letter]] = new_value

    df.to_excel(file_path, index=False)



def add_cumulative_weights():
    ############################################## Cumulative kg Consumed ################################################
    import pandas as pd
    df2 = pd.read_excel(feeding_rate_path)

    # Iterate through the rows and columns of the DataFrame
    for index, row in df2.iterrows():
        for col in df2.columns:
            cell_value = row[col]
            if cell_value == "Cumulative kg consumed":
                # Print the row, column, and cell coordinates
                target_row = index + 1
                target_col = int(col[-1])
                break

    # Specify the column and the starting row
    column_index = target_col
    starting_row = target_row

    # All values under "Cumulative kg consumed" column in a list
    kg_values = df2.iloc[starting_row:, column_index].tolist()

    ###################################################### Return Time Feed ##################################################
    # Read the Excel file into a DataFrame
    df = pd.read_excel(feeding_rate_path)

    # Find the column index for the "Timestamp" header
    for index, row in df.iterrows():
        for col_index, col in enumerate(df.columns):
            cell_value = row[col_index]
            if cell_value == "Timestamp (hh:mm format)":
                # Print the row, column, and cell coordinates
                target_row = index + 1
                target_col = col_index
                break

    # Specify the column and the starting row
    column_index = target_col
    starting_row = target_row

    # All values under "Cumulative kg consumed" column in a list
    values = df.iloc[starting_row:, column_index].tolist()
    time_feed = []

    for i in values:
        try:
            formatted_time = i.strftime("%H:%M")
            time_feed.append(formatted_time)
        except:
            continue

    # return time_feed
    print(time_feed)

    # Create a dictionary with time as keys and kg as values
    data_dict = dict(zip(time_feed, kg_values))

    ################################################ Return Time N #################################################################

    import pandas as pd
    from datetime import datetime

    # Read the Excel file into a DataFrame
    df = pd.read_excel(file_path)

    # Specify the column name containing the timestamp data
    timestamp_column = 'Time'

    # Extract the timestamp data column
    timestamp_data = df[timestamp_column].tolist()

    # Print the timestamp data
    time_N = []
    for timestamp in timestamp_data:
        datetime_obj = datetime.strptime(timestamp, '%I:%M:%S %p')
        time_str = datetime_obj.strftime('%H:%M:%S')
        time_N.append(time_str)

    ############################################### Time Difference ###############################################################

    import pandas as pd
    from datetime import datetime

    # Define the two time points
    # time1 = datetime.strptime('10:30:00', '%H:%M:%S')
    # time2 = datetime.strptime('14:45:30', '%H:%M:%S')

    df = pd.read_excel(file_path)

    # Find the time difference for each pair of times
    time_feed_new = [datetime.strptime(time, "%H:%M").strftime("%H:%M:%S") for time in time_feed]

    time_feed_dt = [datetime.strptime(time, "%H:%M:%S") for time in time_feed_new]
    time_N_dt = [datetime.strptime(time, "%H:%M:%S") for time in time_N]

    for i, N_time in enumerate(time_N_dt):
        smallest_diff = None
        smallest_diff_index_feed = None

        # Iterate over each time_feed element
        for j, feed_time in enumerate(time_feed_dt):
            # Calculate time difference
            diff = abs(N_time - feed_time)

            # Update smallest difference and index if it is the smallest encountered so far
            if smallest_diff is None or diff < smallest_diff:
                smallest_diff = diff
                smallest_diff_index_feed = j

        # Get the corresponding time_feed and combustion kg based on the smallest time difference
        closest_feed_time = time_feed[smallest_diff_index_feed]
        cumulative_kg = data_dict[closest_feed_time]

        # Update the corresponding row in the DataFrame with the combustion kg value
        df.at[i, 'Cumulative kg'] = cumulative_kg

    # Save the updated DataFrame to a new Excel file
    df.to_excel(file_path)


def rolling_avg(columns):
    """
    :param columns: <list> columns of data that you want rolling averages for
    :return: Edited Excel Sheet such that the data in certain columns of the Excel sheet are now rolling averages
    """

    # Load the Excel file
    workbook = load_workbook(file_path)
    sheet = workbook['Sheet1']  # Replace 'Sheet1' with the actual sheet name

    # Define the column letters
    column_letters = []  # Replace with your actual list of column letters

    for i in columns:
        column_letters.append(col_letters[i])

    # Iterate through the columns
    for column_letter in column_letters:
        column = sheet[column_letter]  # Get the column using the letter
        temp_list = []
        for cell in column:
            cell_value = cell.value
            if cell > 5:
                # Assign a new value to the cell
                new_value = sum(temp_list) / 5  # Replace with your desired new value
                cell.value = new_value
                temp_list.pop(0)
            else:
                temp_list.append(cell_value)

    # Save the modified Excel file
    workbook.save(file_path)

def rid_extremes(columns):
    """
    :param columns: Columns in Excel that you want to get rid of extremes in
    :return: Edited Excel sheet with those columns changed such that extremes are removed
    """
    high_extreme = 100 # Change according to preference
    workbook = load_workbook(file_path)
    sheet = workbook['Sheet1']  # Replace 'Sheet1' with the actual sheet name

    # Define the column letters
    column_letters = []  # Replace with your actual list of column letters

    for i in columns:
        column_letters.append(col_letters[i])

    # Iterate through the columns
    for column_letter in column_letters:
        column = sheet[column_letter]  # Get the column using the letter
        for cell in column:
            cell_value = cell.value
            if cell_value > high_extreme:
                # Assign a new value to the cell
                new_value = 0 # Replace with your desired new value
                cell.value = new_value

    # Save the modified Excel file
    workbook.save(file_path)

# Running Data

fan_data(fan_cols)
add_cumulative_weights()